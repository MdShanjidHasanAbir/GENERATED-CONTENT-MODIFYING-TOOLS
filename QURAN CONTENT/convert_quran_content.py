from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SUPPORTED_JSON_STRUCTURES = ("juz.json", "surah.json", "page.json")
SHEET_DATA_FONT_SIZE = 14
SHEET_ROW_HEIGHT = 46
SHEET_COLUMN_WIDTH = 25
HEADER_FILL_COLOR = "FCE4D6"
DATA_FONT = Font(size=SHEET_DATA_FONT_SIZE)
HEADER_FONT = Font(size=SHEET_DATA_FONT_SIZE, bold=True)
HEADER_FILL = PatternFill(fill_type="solid", fgColor=HEADER_FILL_COLOR)
CELL_ALIGNMENT = Alignment(vertical="center", wrap_text=False)


@dataclass(frozen=True)
class QuranConversionSummary:
    input_file: str
    output_file: str
    sheet_count: int
    row_count: int


@dataclass(frozen=True)
class QuranJsonUpdateSummary:
    input_file: str
    output_file: str
    sheet_count: int
    row_count: int


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return " ".join(text.replace("\u00a0", " ").split())


def normalize_id(value) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return text
    if number == number.to_integral_value():
        return str(number.quantize(Decimal(1)))
    return format(number.normalize(), "f")


def normalize_quran_content_json(value, structure_type: str) -> str:
    if structure_type not in SUPPORTED_JSON_STRUCTURES:
        raise ValueError(f"Unsupported Quran JSON structure: {structure_type}")

    source = _source_content_data(value)
    converted = {
        "meta": {
            "title": normalize_text(_nested_value(source, ("meta", "title"))),
            "description": normalize_text(_nested_value(source, ("meta", "description"))),
        },
        "heading": {
            "title": normalize_text(_heading_title(source)),
            "description": normalize_text(_heading_description(source)),
        },
    }

    if structure_type == "surah.json":
        converted["lessons"] = _string_list(source.get("lessons"))
        converted["faqs"] = _faq_list(source.get("faqs"))
    else:
        converted["searching_terms"] = _string_list(source.get("searching_terms"))

    return json.dumps(converted, ensure_ascii=False, separators=(",", ":"))


def _source_content_data(value) -> dict:
    text = normalize_text(value)
    data = _parse_json_value(text)
    if data is None:
        return {"heading": {"description": text}}
    if isinstance(data, list):
        data = next((item for item in data if isinstance(item, dict)), {})
    if isinstance(data, dict) and isinstance(data.get("result"), dict):
        data = data["result"]
    return data if isinstance(data, dict) else {"heading": {"description": text}}


def _parse_json_value(text: str):
    if not text:
        return None
    candidates = [text]
    for marker in ("{", "["):
        index = text.find(marker)
        if index > 0:
            candidates.append(text[index:])

    decoder = json.JSONDecoder()
    for candidate in candidates:
        for parser in (json.loads, lambda raw: decoder.raw_decode(raw)[0]):
            try:
                return parser(candidate)
            except json.JSONDecodeError:
                continue
    return None


def _nested_value(data: dict, keys: tuple[str, ...]):
    current = data
    for key in keys:
        if not isinstance(current, dict):
            return None
        current = current.get(key)
    return current


def _heading_title(data: dict):
    heading = data.get("heading")
    if isinstance(heading, dict):
        return heading.get("title")
    return heading or data.get("title")


def _heading_description(data: dict):
    heading = data.get("heading")
    if isinstance(heading, dict) and heading.get("description") is not None:
        return heading.get("description")
    return data.get("summary") or data.get("description")


def _string_list(value) -> list[str]:
    if isinstance(value, list):
        return [normalize_text(item) for item in value if normalize_text(item)]
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(",") if part.strip()]


def _faq_list(value) -> list[dict]:
    if not isinstance(value, list):
        return []
    faqs = []
    for item in value:
        if not isinstance(item, dict):
            continue
        question = normalize_text(item.get("question"))
        answer = normalize_text(item.get("answer"))
        if question or answer:
            faqs.append({"question": question, "answer": answer})
    return faqs


def quran_id_column(path: Path | str) -> str:
    name = normalize_text(Path(path).stem).casefold()
    if "juz" in name:
        return "juz_id"
    if "page" in name:
        return "page_id"
    if "surah" in name:
        return "surah_id"
    return "id"


def language_id(path: Path | str) -> str:
    name = normalize_text(Path(path).name)
    match = re.match(r"([A-Za-z]{2})(?:[_\-\s]|$)", name)
    if match:
        return match.group(1).lower()
    return ""


def discover_workbooks(directory: Path | str) -> list[Path]:
    return sorted(path for path in Path(directory).rglob("*.xlsx") if not path.name.startswith("~$"))


def available_quran_output_workbooks(output_dir: Path | str = Path("OUTPUT")) -> list[Path]:
    return discover_workbooks(output_dir)


def convert_quran_content_folder(
    input_dir: Path | str = Path("INPUT"),
    output_dir: Path | str = Path("OUTPUT"),
) -> list[QuranConversionSummary]:
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    summaries = []
    for input_path in discover_workbooks(input_dir):
        output_path = output_dir / input_path.name
        summaries.append(convert_quran_content_workbook(input_path, output_path))

    write_summary_csv(output_dir / "conversion_summary.csv", summaries)
    return summaries


def convert_quran_content_workbook(input_path: Path | str, output_path: Path | str) -> QuranConversionSummary:
    input_path = Path(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    id_header = quran_id_column(input_path)
    lang = language_id(input_path)
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    output_workbook = Workbook(write_only=True)
    row_count = 0
    sheet_count = 0
    try:
        for sheet in workbook.worksheets:
            output_sheet = output_workbook.create_sheet(title=sheet.title)
            output_sheet.append([id_header, "language_id", "contents"])
            sheet_count += 1

            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            id_index = columns.get("id")
            response_index = columns.get("response")
            if id_index is None or response_index is None:
                continue

            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                row_id = normalize_id(_cell(row, id_index))
                contents = normalize_text(_cell(row, response_index))
                if not row_id or not contents:
                    continue
                output_sheet.append([row_id, lang, contents])
                row_count += 1

        output_workbook.save(output_path)
    finally:
        workbook.close()

    return QuranConversionSummary(
        input_file=input_path.name,
        output_file=output_path.name,
        sheet_count=sheet_count,
        row_count=row_count,
    )


def update_quran_output_files(
    output_dir: Path | str = Path("OUTPUT"),
    updated_dir: Path | str = Path("UPDATED OUTPUT"),
    structure_type: str = "juz.json",
    filenames: Iterable[str] | None = None,
) -> list[QuranJsonUpdateSummary]:
    output_dir = Path(output_dir)
    updated_dir = Path(updated_dir)
    if structure_type not in SUPPORTED_JSON_STRUCTURES:
        raise ValueError(f"Unsupported Quran JSON structure: {structure_type}")

    if filenames:
        input_paths = []
        for filename in filenames:
            input_path = output_dir / filename
            if not input_path.exists():
                raise FileNotFoundError(f"Selected Quran output file not found: {filename}")
            input_paths.append(input_path)
    else:
        input_paths = available_quran_output_workbooks(output_dir)

    summaries = []
    for input_path in input_paths:
        output_path = updated_dir / input_path.relative_to(output_dir)
        summaries.append(update_quran_output_workbook(input_path, output_path, structure_type))
    return summaries


def update_quran_output_workbook(
    input_path: Path | str,
    output_path: Path | str,
    structure_type: str,
) -> QuranJsonUpdateSummary:
    input_path = Path(input_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(input_path, read_only=True, data_only=True)
    output_workbook = Workbook(write_only=True)
    row_count = 0
    sheet_count = 0
    try:
        for sheet in workbook.worksheets:
            output_sheet = output_workbook.create_sheet(title=sheet.title)
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            header_list = list(header)
            columns = _header_map(header)
            content_index = columns.get("contents")
            _format_output_sheet_columns(output_sheet, len(header_list))
            row_number = 1
            _append_formatted_row(output_sheet, header_list, row_number, is_header=True)
            row_number += 1
            sheet_count += 1

            for row in rows:
                output_row = list(row)
                if content_index is not None and content_index < len(output_row) and output_row[content_index]:
                    output_row[content_index] = normalize_quran_content_json(output_row[content_index], structure_type)
                    row_count += 1
                _append_formatted_row(output_sheet, output_row, row_number)
                row_number += 1
        output_workbook.save(output_path)
        apply_output_workbook_format(output_path)
    finally:
        workbook.close()

    return QuranJsonUpdateSummary(
        input_file=input_path.name,
        output_file=output_path.name,
        sheet_count=sheet_count,
        row_count=row_count,
    )


def write_summary_csv(path: Path | str, summaries: Iterable[QuranConversionSummary]) -> None:
    path = Path(path)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=["input_file", "output_file", "sheet_count", "row_count"])
        writer.writeheader()
        for summary in summaries:
            writer.writerow(
                {
                    "input_file": summary.input_file,
                    "output_file": summary.output_file,
                    "sheet_count": summary.sheet_count,
                    "row_count": summary.row_count,
                }
            )


def print_summary(summaries: Iterable[QuranConversionSummary]) -> None:
    for summary in summaries:
        print(
            f"{summary.input_file} -> {summary.output_file}: "
            f"sheets={summary.sheet_count}, rows={summary.row_count}"
        )


def print_update_summary(summaries: Iterable[QuranJsonUpdateSummary]) -> None:
    for summary in summaries:
        print(
            f"{summary.input_file} -> {summary.output_file}: "
            f"sheets={summary.sheet_count}, updated_rows={summary.row_count}"
        )


def _header_map(header) -> dict[str, int]:
    if not header:
        return {}
    return {normalize_text(value).casefold(): index for index, value in enumerate(header) if normalize_text(value)}


def _cell(row, index: int):
    return row[index] if index < len(row) else None


def apply_output_workbook_format(path: Path | str) -> None:
    workbook = load_workbook(path)
    try:
        for sheet in workbook.worksheets:
            _format_existing_sheet(sheet)
        workbook.save(path)
    finally:
        workbook.close()


def _format_existing_sheet(sheet) -> None:
    for column_index in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = SHEET_COLUMN_WIDTH
    for row_index in range(1, sheet.max_row + 1):
        sheet.row_dimensions[row_index].height = SHEET_ROW_HEIGHT
    for row in sheet.iter_rows():
        for cell in row:
            cell.font = HEADER_FONT if cell.row == 1 else DATA_FONT
            cell.alignment = CELL_ALIGNMENT
            if cell.row == 1:
                cell.fill = HEADER_FILL


def _format_output_sheet_columns(sheet, column_count: int) -> None:
    for column_index in range(1, column_count + 1):
        sheet.column_dimensions[get_column_letter(column_index)].width = SHEET_COLUMN_WIDTH


def _append_formatted_row(sheet, values, row_number: int, is_header: bool = False) -> None:
    sheet.row_dimensions[row_number].height = SHEET_ROW_HEIGHT
    font = HEADER_FONT if is_header else DATA_FONT
    row = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        cell.font = font
        cell.alignment = CELL_ALIGNMENT
        if is_header:
            cell.fill = HEADER_FILL
        row.append(cell)
    sheet.append(row)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Convert and update Quran content workbooks.")
    subparsers = parser.add_subparsers(dest="command")

    convert_parser = subparsers.add_parser("convert", help="Convert input workbooks to id/language_id/contents columns.")
    convert_parser.add_argument("--input-dir", type=Path, default=Path("INPUT"))
    convert_parser.add_argument("--output-dir", type=Path, default=Path("OUTPUT"))

    update_parser = subparsers.add_parser("update-output-json", help="Rewrite OUTPUT workbook contents JSON.")
    update_parser.add_argument("--output-dir", type=Path, default=Path("OUTPUT"))
    update_parser.add_argument("--updated-dir", type=Path, default=Path("UPDATED OUTPUT"))
    update_parser.add_argument("--structure", choices=SUPPORTED_JSON_STRUCTURES, required=True)
    update_parser.add_argument("--files", nargs="*", default=None)

    parser.add_argument("--input-dir", type=Path, default=Path("INPUT"))
    parser.add_argument("--output-dir", type=Path, default=Path("OUTPUT"))
    args = parser.parse_args(argv)

    if args.command == "update-output-json":
        try:
            summaries = update_quran_output_files(args.output_dir, args.updated_dir, args.structure, args.files)
        except (FileNotFoundError, ValueError) as exc:
            print(str(exc), file=sys.stderr)
            return 1
        print_update_summary(summaries)
        if not summaries:
            print("No Quran output workbooks found.", file=sys.stderr)
            return 1
        print(f"Updated Quran content workbooks written to: {args.updated_dir}")
        return 0

    summaries = convert_quran_content_folder(args.input_dir, args.output_dir)
    print_summary(summaries)

    if not summaries:
        print("No Quran content workbooks found.", file=sys.stderr)
        return 1
    print(f"Converted Quran content workbooks written to: {args.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
