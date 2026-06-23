import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / "convert_quran_content.py"
spec = importlib.util.spec_from_file_location("convert_quran_content", MODULE_PATH)
quran = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = quran
spec.loader.exec_module(quran)


class ConvertQuranContentTest(unittest.TestCase):
    def test_infers_output_id_column_from_quran_content_filename(self):
        self.assertEqual(quran.quran_id_column(Path("EN_Quran Juz Content.xlsx")), "juz_id")
        self.assertEqual(quran.quran_id_column(Path("EN_Quran Page Content.xlsx")), "page_id")
        self.assertEqual(quran.quran_id_column(Path("EN_Quran Surah Content.xlsx")), "surah_id")

    def test_converts_workbook_to_requested_three_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "EN_Quran Surah Content.xlsx"
            output_path = root / "OUTPUT" / input_path.name
            self._write_source_workbook(input_path, rows=[["1", "chatgpt", "id: 1", "Yes", "time", "response one"]])

            summary = quran.convert_quran_content_workbook(input_path, output_path)

            self.assertEqual(summary.input_file, input_path.name)
            self.assertEqual(summary.output_file, output_path.name)
            self.assertEqual(summary.row_count, 1)
            self.assertEqual(
                self._read_rows(output_path),
                [
                    ("surah_id", "language_id", "contents"),
                    ("1", "en", "response one"),
                ],
            )

    def test_converts_every_sheet_and_skips_rows_without_id_or_response(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "BN_Quran Page Content.xlsx"
            output_path = root / "OUTPUT" / input_path.name

            workbook = Workbook()
            first = workbook.active
            first.title = "Results"
            first.append(["ID", "Platform", "Keyword", "Success", "Timestamp", "Response"])
            first.append([2.0, "chatgpt", "id_page: 2", "Yes", "time", "page response"])
            first.append([3, "chatgpt", "id_page: 3", "Yes", "time", None])
            second = workbook.create_sheet(title="Extra")
            second.append(["ID", "Response"])
            second.append([4, "extra response"])
            workbook.save(input_path)

            summary = quran.convert_quran_content_workbook(input_path, output_path)

            self.assertEqual(summary.row_count, 2)
            self.assertEqual(
                self._read_rows(output_path, sheet_name="Results"),
                [
                    ("page_id", "language_id", "contents"),
                    ("2", "bn", "page response"),
                ],
            )
            self.assertEqual(
                self._read_rows(output_path, sheet_name="Extra"),
                [
                    ("page_id", "language_id", "contents"),
                    ("4", "bn", "extra response"),
                ],
            )

    def test_converts_folder_outputs_each_input_workbook(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "INPUT"
            output_dir = root / "OUTPUT"
            input_dir.mkdir()
            self._write_source_workbook(
                input_dir / "EN_Quran Juz Content.xlsx",
                rows=[["1", "chatgpt", "id: 1", "Yes", "time", "juz response"]],
            )
            self._write_source_workbook(
                input_dir / "EN_Quran Surah Content.xlsx",
                rows=[["1", "chatgpt", "id: 1", "Yes", "time", "surah response"]],
            )

            summaries = quran.convert_quran_content_folder(input_dir, output_dir)

            self.assertEqual([summary.output_file for summary in summaries], ["EN_Quran Juz Content.xlsx", "EN_Quran Surah Content.xlsx"])
            self.assertTrue((output_dir / "EN_Quran Juz Content.xlsx").exists())
            self.assertTrue((output_dir / "EN_Quran Surah Content.xlsx").exists())

    def test_normalizes_wrapped_result_to_juz_json_shape(self):
        source = json.dumps(
            [
                {
                    "keyword": "al quran juz 1",
                    "result": {
                        "meta": {"title": "Juz title", "description": "Juz description"},
                        "heading": "Al Quran Juz 1",
                        "summary": "Juz summary",
                        "searching_terms": "term one, term two, term three",
                    },
                }
            ]
        )

        converted = json.loads(quran.normalize_quran_content_json(source, "juz.json"))

        self.assertEqual(set(converted), {"meta", "heading", "searching_terms"})
        self.assertEqual(converted["meta"], {"title": "Juz title", "description": "Juz description"})
        self.assertEqual(converted["heading"], {"title": "Al Quran Juz 1", "description": "Juz summary"})
        self.assertEqual(converted["searching_terms"], ["term one", "term two", "term three"])

    def test_normalizes_direct_surah_json_shape(self):
        source = json.dumps(
            {
                "meta": {"title": "Surah title", "description": "Surah description"},
                "heading": {"title": "Surah Al Fatihah", "description": "Heading description"},
                "lessons": ["lesson one"],
                "faqs": [{"question": "Question?", "answer": "Answer."}],
                "searching_terms": ["ignored for surah"],
            }
        )

        converted = json.loads(quran.normalize_quran_content_json(source, "surah.json"))

        self.assertEqual(set(converted), {"meta", "heading", "lessons", "faqs"})
        self.assertEqual(converted["meta"]["title"], "Surah title")
        self.assertEqual(converted["heading"], {"title": "Surah Al Fatihah", "description": "Heading description"})
        self.assertEqual(converted["lessons"], ["lesson one"])
        self.assertEqual(converted["faqs"], [{"question": "Question?", "answer": "Answer."}])

    def test_normalizes_invalid_json_to_valid_page_shape(self):
        converted = json.loads(quran.normalize_quran_content_json("plain text content", "page.json"))

        self.assertEqual(set(converted), {"meta", "heading", "searching_terms"})
        self.assertEqual(converted["meta"], {"title": "", "description": ""})
        self.assertEqual(converted["heading"], {"title": "", "description": "plain text content"})
        self.assertEqual(converted["searching_terms"], [])

    def test_updates_selected_output_workbooks_to_updated_output_folder(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "OUTPUT"
            updated_dir = root / "UPDATED OUTPUT"
            output_dir.mkdir()
            source_name = "EN_Quran Juz Content.xlsx"
            self._write_converted_output_workbook(
                output_dir / source_name,
                headers=["juz_id", "language_id", "contents"],
                rows=[
                    [
                        "1",
                        "en",
                        json.dumps(
                            [
                                {
                                    "result": {
                                        "meta": {"title": "Juz title", "description": "Juz description"},
                                        "heading": "Al Quran Juz 1",
                                        "summary": "Juz summary",
                                        "searching_terms": "term one, term two",
                                    }
                                }
                            ]
                        ),
                    ]
                ],
            )

            summaries = quran.update_quran_output_files(output_dir, updated_dir, "juz.json", [source_name])

            updated_path = updated_dir / source_name
            rows = self._read_rows(updated_path)
            converted = json.loads(rows[1][2])
            self.assertEqual([summary.output_file for summary in summaries], [source_name])
            self.assertEqual(rows[0], ("juz_id", "language_id", "contents"))
            self.assertEqual(converted["heading"], {"title": "Al Quran Juz 1", "description": "Juz summary"})
            self.assertEqual(converted["searching_terms"], ["term one", "term two"])
            self._assert_formatted_sheet(updated_path, expected_columns=3)

    def test_update_output_files_reports_missing_selected_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "OUTPUT"
            updated_dir = root / "UPDATED OUTPUT"
            output_dir.mkdir()

            with self.assertRaises(FileNotFoundError) as context:
                quran.update_quran_output_files(output_dir, updated_dir, "surah.json", ["missing.xlsx"])

            self.assertIn("missing.xlsx", str(context.exception))

    def test_existing_updated_workbook_can_be_formatted_in_place(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "plain.xlsx"
            self._write_converted_output_workbook(
                path,
                headers=["juz_id", "language_id", "contents"],
                rows=[["1", "en", "content"]],
            )

            quran.apply_output_workbook_format(path)

            self._assert_formatted_sheet(path, expected_columns=3)

    def _write_source_workbook(self, path, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(["ID", "Platform", "Keyword", "Success", "Timestamp", "Response"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def _write_converted_output_workbook(self, path, headers, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def _read_rows(self, path, sheet_name="Results"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name]
            return [tuple(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    def _assert_formatted_sheet(self, path, sheet_name="Results", expected_columns=3):
        workbook = load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            self.assertEqual(sheet["A1"].font.sz, 14)
            self.assertTrue(sheet["A1"].font.bold)
            self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00FCE4D6")
            self.assertEqual(sheet["A1"].alignment.vertical, "center")
            self.assertFalse(sheet["A1"].alignment.wrap_text)
            self.assertEqual(sheet["A2"].font.sz, 14)
            self.assertEqual(sheet["A2"].alignment.vertical, "center")
            self.assertFalse(sheet["A2"].alignment.wrap_text)
            self.assertEqual(sheet.row_dimensions[1].height, 46)
            self.assertEqual(sheet.row_dimensions[2].height, 46)
            for index in range(1, expected_columns + 1):
                column_letter = sheet.cell(row=1, column=index).column_letter
                self.assertEqual(sheet.column_dimensions[column_letter].width, 25)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
