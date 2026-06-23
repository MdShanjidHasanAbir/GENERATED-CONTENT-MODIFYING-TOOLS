import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


MODULE_PATH = Path(__file__).resolve().parents[1] / "convert_dua_content.py"
spec = importlib.util.spec_from_file_location("convert_dua_content", MODULE_PATH)
dua = importlib.util.module_from_spec(spec)
sys.modules[spec.name] = dua
spec.loader.exec_module(dua)


class ConvertDuaContentTest(unittest.TestCase):
    def test_converts_workbook_to_id_language_id_contents_columns(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "EN _SINGLE DUA.xlsx"
            output_path = root / "OUTPUT" / input_path.name
            self._write_source_workbook(input_path, rows=[["1", "chatgpt", "keyword", "Yes", "time", "response one"]])

            summary = dua.convert_dua_content_workbook(input_path, output_path)

            self.assertEqual(summary.input_file, input_path.name)
            self.assertEqual(summary.output_file, output_path.name)
            self.assertEqual(summary.row_count, 1)
            self.assertEqual(
                self._read_rows(output_path),
                [
                    ("id", "language_id", "contents"),
                    ("1", "en", "response one"),
                ],
            )

    def test_converts_every_sheet_and_skips_rows_without_id_or_response(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_path = root / "BN_SINGLE DUA.xlsx"
            output_path = root / "OUTPUT" / input_path.name

            workbook = Workbook()
            first = workbook.active
            first.title = "Results"
            first.append(["ID", "Platform", "Keyword", "Success", "Timestamp", "Response"])
            first.append([2.0, "chatgpt", "keyword", "Yes", "time", "dua response"])
            first.append([3, "chatgpt", "keyword", "Yes", "time", None])
            second = workbook.create_sheet(title="Extra")
            second.append(["ID", "Response"])
            second.append([4, "extra response"])
            workbook.save(input_path)

            summary = dua.convert_dua_content_workbook(input_path, output_path)

            self.assertEqual(summary.row_count, 2)
            self.assertEqual(
                self._read_rows(output_path, sheet_name="Results"),
                [
                    ("id", "language_id", "contents"),
                    ("2", "bn", "dua response"),
                ],
            )
            self.assertEqual(
                self._read_rows(output_path, sheet_name="Extra"),
                [
                    ("id", "language_id", "contents"),
                    ("4", "bn", "extra response"),
                ],
            )

    def test_normalizes_wrapped_result_to_dua_json_shape(self):
        source = json.dumps(
            [
                {
                    "keyword": "Need Allah Help",
                    "result": {
                        "meta": {"title": "Meta title", "description": "Meta description"},
                        "heading": {"title": "Heading title"},
                        "description": "Top description",
                        "faqs": [{"question": "Question?", "answer": "Answer."}],
                    },
                }
            ]
        )

        converted = json.loads(dua.normalize_dua_content_json(source))

        self.assertEqual(set(converted), {"meta", "heading", "faqs"})
        self.assertEqual(converted["meta"], {"title": "Meta title", "description": "Meta description"})
        self.assertEqual(converted["heading"], {"title": "Heading title", "description": "Top description"})
        self.assertEqual(converted["faqs"], [{"question": "Question?", "answer": "Answer."}])

    def test_normalizes_direct_json_and_invalid_json(self):
        direct = json.dumps(
            {
                "meta": {"title": "Direct title", "description": "Direct description"},
                "heading": "String heading",
                "summary": "Summary text",
            }
        )

        converted = json.loads(dua.normalize_dua_content_json(direct))
        invalid = json.loads(dua.normalize_dua_content_json("plain text"))

        self.assertEqual(converted["heading"], {"title": "String heading", "description": "Summary text"})
        self.assertEqual(converted["faqs"], [])
        self.assertEqual(invalid["heading"], {"title": "", "description": "plain text"})

    def test_updates_selected_output_workbooks_to_updated_content_folder(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "OUTPUT"
            updated_dir = root / "UPDATED CONTENT"
            output_dir.mkdir()
            source_name = "EN _SINGLE DUA.xlsx"
            self._write_converted_output_workbook(
                output_dir / source_name,
                headers=["id", "language_id", "contents"],
                rows=[
                    [
                        "1",
                        "en",
                        json.dumps(
                            [
                                {
                                    "result": {
                                        "meta": {"title": "Meta title", "description": "Meta description"},
                                        "heading": "Need Allah Help",
                                        "description": "Description text",
                                        "faqs": [{"question": "Q?", "answer": "A."}],
                                    }
                                }
                            ]
                        ),
                    ]
                ],
            )

            summaries = dua.update_dua_output_files(output_dir, updated_dir, [source_name])

            rows = self._read_rows(updated_dir / source_name)
            converted = json.loads(rows[1][2])
            self.assertEqual([summary.output_file for summary in summaries], [source_name])
            self.assertEqual(rows[0], ("id", "language_id", "contents"))
            self.assertEqual(converted["heading"], {"title": "Need Allah Help", "description": "Description text"})
            self._assert_formatted_sheet(updated_dir / source_name, expected_columns=3)

    def test_update_output_files_reports_missing_selected_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            output_dir = root / "OUTPUT"
            updated_dir = root / "UPDATED CONTENT"
            output_dir.mkdir()

            with self.assertRaises(FileNotFoundError) as context:
                dua.update_dua_output_files(output_dir, updated_dir, ["missing.xlsx"])

            self.assertIn("missing.xlsx", str(context.exception))

    def test_existing_updated_workbook_can_be_formatted_in_place(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "plain.xlsx"
            self._write_converted_output_workbook(
                path,
                headers=["id", "language_id", "contents"],
                rows=[["1", "en", "content"]],
            )

            dua.apply_output_workbook_format(path)

            self._assert_formatted_sheet(path, expected_columns=3)

    def _write_source_workbook(self, path, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(["ID", "Platform", "Keyword", "Success", "Timestamp", "Response"])
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def _write_converted_output_workbook(self, path, headers, rows):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Results"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)

    def _read_rows(self, path, sheet_name="Results"):
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name]
            return [tuple(row) for row in sheet.iter_rows(values_only=True)]
        finally:
            workbook.close()

    def _assert_formatted_sheet(self, path, sheet_name="Results", expected_columns=3):
        workbook = load_workbook(path)
        try:
            sheet = workbook[sheet_name]
            self.assertEqual(sheet["A1"].font.sz, 14)
            self.assertTrue(sheet["A1"].font.bold)
            self.assertEqual(sheet["A1"].fill.fgColor.rgb, "00FCE4D6")
            self.assertEqual(sheet["A1"].alignment.vertical, "center")
            self.assertFalse(sheet["A1"].alignment.wrap_text)
            self.assertEqual(sheet["A2"].font.sz, 14)
            self.assertEqual(sheet["A2"].alignment.vertical, "center")
            self.assertFalse(sheet["A2"].alignment.wrap_text)
            self.assertEqual(sheet.row_dimensions[1].height, 46)
            self.assertEqual(sheet.row_dimensions[2].height, 46)
            for index in range(1, expected_columns + 1):
                column_letter = sheet.cell(row=1, column=index).column_letter
                self.assertEqual(sheet.column_dimensions[column_letter].width, 25)
        finally:
            workbook.close()


if __name__ == "__main__":
    unittest.main()
