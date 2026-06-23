import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


def load_hadith_module():
    module_path = Path(__file__).resolve().parents[1] / "SINGLE HADITH CONTENT" / "reconcile_hadith_outputs.py"
    spec = importlib.util.spec_from_file_location("reconcile_hadith_outputs_for_test", module_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


hadith = load_hadith_module()


def write_workbook(path: Path, header: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def read_rows(path: Path) -> list[tuple]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


class HadithSelectiveParsingTest(unittest.TestCase):
    def test_related_hadith_object_ids_are_converted_to_ascii_digits(self):
        content = json.dumps(
            {
                "heading": "Example",
                "related_hadiths": [
                    {"id": "21\u09ef", "book_id": "1", "label": "Sahih Bukhari"},
                    {"id": "52\u06f8a", "book_id": "6", "label": "Sunan Ibn Majah"},
                ],
            },
            ensure_ascii=False,
        )

        converted, errors = hadith.convert_final_content_json_with_errors(content, language="en")

        related = json.loads(converted)["related_hadiths"]
        self.assertEqual(errors, "")
        self.assertEqual(
            related,
            [
                {"id": "219", "book_id": "1", "label": "Sahih Bukhari"},
                {"id": "528a", "book_id": "6", "label": "Sunan Ibn Majah"},
            ],
        )

    def test_related_hadith_text_uses_books_catalog_for_available_extra_books(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            books_path = root / "BOOKS.xlsx"
            final_dir = root / "BOOK WISE FINAL"
            updated_dir = root / "BOOK WISE FINAL UPDATED CONTENT"
            write_workbook(
                books_path,
                ["id", "language_id", "book_name"],
                [["88", "en", "ADABUL MUFRAD"], ["99", "en", "100 HADITH"]],
            )
            content = json.dumps(
                {
                    "heading": "Example",
                    "related_hadiths": ["Adabul Mufrad 45", "100 Hadith \u09ed"],
                },
                ensure_ascii=False,
            )
            write_workbook(
                final_dir / "EN" / "EN_EXTRA.xlsx",
                ["book_id", "language_id", "hadith_id", "content"],
                [["88", "en", "1", content]],
            )

            hadith.write_updated_content_workbooks(final_dir, updated_dir, books_path)

            rows = read_rows(updated_dir / "EN" / "EN_EXTRA.xlsx")
            converted = json.loads(rows[1][3])
            self.assertEqual(
                converted["related_hadiths"],
                [
                    {"id": "45", "book_id": "88", "label": "Adabul Mufrad"},
                    {"id": "7", "book_id": "99", "label": "100 Hadith"},
                ],
            )
            self.assertEqual(rows[0], ("book_id", "language_id", "hadith_id", "content", "content_error_details"))
            self.assertTrue(len(rows[1]) < 5 or rows[1][4] is None)

    def test_updated_content_converts_mixed_bengali_related_hadith_ids_to_ascii(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            books_path = root / "BOOKS.xlsx"
            final_dir = root / "BOOK WISE FINAL"
            updated_dir = root / "BOOK WISE FINAL UPDATED CONTENT"
            write_workbook(
                books_path,
                ["id", "language_id", "book_name"],
                [["1", "bn", "BUKHARI"], ["2", "bn", "MUSLIM"]],
            )
            content = json.dumps(
                {
                    "heading": "Example",
                    "related_hadiths": ["Sahih Bukhari 631\u09ee", "Sahih Muslim 1\u09ef0\u09ed"],
                },
                ensure_ascii=False,
            )
            write_workbook(
                final_dir / "BN" / "BN_BUKHARI.xlsx",
                ["book_id", "language_id", "hadith_id", "content"],
                [["1", "bn", "1", content]],
            )

            hadith.write_updated_content_workbooks(final_dir, updated_dir, books_path)

            rows = read_rows(updated_dir / "BN" / "BN_BUKHARI.xlsx")
            converted = json.loads(rows[1][3])
            self.assertEqual(
                converted["related_hadiths"],
                [
                    {"id": "6318", "book_id": "1", "label": "Sahih Bukhari"},
                    {"id": "1907", "book_id": "2", "label": "Sahih Muslim"},
                ],
            )
            self.assertTrue(len(rows[1]) < 5 or rows[1][4] is None)

    def test_selected_reconciled_files_update_book_wise_final_without_removing_existing_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            books_path = root / "BOOKS.xlsx"
            reconciled_dir = root / "reconciled_output"
            final_dir = root / "BOOK WISE FINAL"
            write_workbook(
                books_path,
                ["id", "language_id", "book_name"],
                [["101", "en", "BUKHARI"], ["102", "en", "MUSLIM"]],
            )
            write_workbook(
                reconciled_dir / "EN_BUKHARI_MARGE.xlsx",
                ["id", "response"],
                [["1", "{\"heading\":\"bukhari\"}"]],
            )
            write_workbook(
                reconciled_dir / "EN_MUSLIM_MARGE.xlsx",
                ["id", "response"],
                [["2", "{\"heading\":\"muslim\"}"]],
            )
            write_workbook(
                final_dir / "EN" / "EN_MUSLIM.xlsx",
                ["book_id", "language_id", "hadith_id", "content"],
                [["existing", "en", "existing", "keep me"]],
            )

            counts = hadith.write_book_wise_final(
                reconciled_dir,
                books_path,
                final_dir,
                source_paths=[reconciled_dir / "EN_BUKHARI_MARGE.xlsx"],
                clear_existing=False,
            )

            self.assertEqual(counts, {"EN_BUKHARI": 1})
            self.assertEqual(
                read_rows(final_dir / "EN" / "EN_BUKHARI.xlsx"),
                [
                    ("book_id", "language_id", "hadith_id", "content"),
                    ("101", "en", "1", "{\"heading\":\"bukhari\"}"),
                ],
            )
            assert_formatted_sheet(self, final_dir / "EN" / "EN_BUKHARI.xlsx", expected_columns=4)
            self.assertEqual(
                read_rows(final_dir / "EN" / "EN_MUSLIM.xlsx"),
                [
                    ("book_id", "language_id", "hadith_id", "content"),
                    ("existing", "en", "existing", "keep me"),
                ],
            )

    def test_selected_final_files_update_content_without_removing_existing_files(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            books_path = root / "BOOKS.xlsx"
            final_dir = root / "BOOK WISE FINAL"
            updated_dir = root / "BOOK WISE FINAL UPDATED CONTENT"
            write_workbook(
                books_path,
                ["id", "language_id", "book_name"],
                [["101", "en", "BUKHARI"], ["102", "en", "MUSLIM"]],
            )
            write_workbook(
                final_dir / "EN" / "EN_BUKHARI.xlsx",
                ["book_id", "language_id", "hadith_id", "content"],
                [["101", "en", "1", "plain content"]],
            )
            write_workbook(
                updated_dir / "EN" / "EN_MUSLIM.xlsx",
                ["book_id", "language_id", "hadith_id", "content"],
                [["existing", "en", "existing", "keep me"]],
            )

            counts = hadith.write_updated_content_workbooks(
                final_dir,
                updated_dir,
                books_path,
                source_paths=[final_dir / "EN" / "EN_BUKHARI.xlsx"],
                clear_existing=False,
            )

            self.assertEqual(counts, {"EN/EN_BUKHARI.xlsx": 1})
            rows = read_rows(updated_dir / "EN" / "EN_BUKHARI.xlsx")
            converted = json.loads(rows[1][3])
            self.assertEqual(converted["heading"]["description"], "plain content")
            self.assertEqual(rows[0], ("book_id", "language_id", "hadith_id", "content", "content_error_details"))
            assert_formatted_sheet(self, updated_dir / "EN" / "EN_BUKHARI.xlsx", expected_columns=5)
            self.assertEqual(
                read_rows(updated_dir / "EN" / "EN_MUSLIM.xlsx"),
                [
                    ("book_id", "language_id", "hadith_id", "content"),
                    ("existing", "en", "existing", "keep me"),
                ],
            )

    def test_existing_output_workbook_can_be_formatted_in_place(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = Path(tmp) / "plain.xlsx"
            write_workbook(
                path,
                ["book_id", "language_id", "hadith_id", "content"],
                [["1", "en", "10", "content"]],
            )

            hadith.apply_output_workbook_format(path)

            assert_formatted_sheet(self, path, expected_columns=4)


def assert_formatted_sheet(test_case: unittest.TestCase, path: Path, expected_columns: int) -> None:
    workbook = load_workbook(path)
    try:
        sheet = workbook.active
        test_case.assertEqual(sheet["A1"].font.sz, 14)
        test_case.assertTrue(sheet["A1"].font.bold)
        test_case.assertEqual(sheet["A1"].fill.fgColor.rgb, "00FCE4D6")
        test_case.assertEqual(sheet["A1"].alignment.vertical, "center")
        test_case.assertFalse(sheet["A1"].alignment.wrap_text)
        test_case.assertEqual(sheet["A2"].font.sz, 14)
        test_case.assertEqual(sheet["A2"].alignment.vertical, "center")
        test_case.assertFalse(sheet["A2"].alignment.wrap_text)
        test_case.assertEqual(sheet.row_dimensions[1].height, 46)
        test_case.assertEqual(sheet.row_dimensions[2].height, 46)
        for index in range(1, expected_columns + 1):
            column_letter = sheet.cell(row=1, column=index).column_letter
            test_case.assertEqual(sheet.column_dimensions[column_letter].width, 25)
    finally:
        workbook.close()


if __name__ == "__main__":
    unittest.main()
