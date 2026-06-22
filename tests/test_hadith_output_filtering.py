import importlib.util
import json
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


def load_hadith_module():
    module_path = Path(__file__).resolve().parents[1] / "SINGLE HADITH CONTENT" / "reconcile_hadith_outputs.py"
    spec = importlib.util.spec_from_file_location("reconcile_hadith_outputs_filter_test", module_path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    spec.loader.exec_module(module)
    return module


hadith = load_hadith_module()


def write_workbook(path: Path, header: list[str], rows: list[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(header)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def read_rows(path: Path) -> list[tuple]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return [tuple(row) for row in workbook.active.iter_rows(values_only=True)]
    finally:
        workbook.close()


def keyword(hadith_id: str, translation: str = "বাংলা অনুবাদ", arabic: str = "") -> str:
    parts = [f"id: {hadith_id}"]
    if arabic:
        parts.append(f"arabic: {arabic}")
    parts.append(f"translation: {translation}")
    return "\n".join(parts)


def response(text: str, nested: bool = True) -> str:
    if nested:
        return json.dumps({"content": {"heading": text, "explanation": text}}, ensure_ascii=False)
    return json.dumps({"heading": text, "explanation": text}, ensure_ascii=False)


class HadithOutputFilteringTest(unittest.TestCase):
    def test_blocked_response_terms_are_removed_and_added_to_missing_input(self):
        with tempfile.TemporaryDirectory() as tmp:
            root = Path(tmp)
            input_dir = root / "INPUT"
            output_dir = root / "OUTPUT"
            cleaned_dir = root / "reconciled_output"
            missing_dir = root / "missing_input"
            input_path = input_dir / "BN" / "[V1] BUKHARI-BANGLA.xlsx"
            output_path = output_dir / "BN_BUKHARI.xlsx"
            write_workbook(
                input_path,
                ["id", "arabic", "translation"],
                [
                    ["1", "", "বাংলা এক"],
                    ["2", "", "বাংলা দুই"],
                    ["3", "", "বাংলা তিন"],
                    ["4", "", "বাংলা চার"],
                    ["5", "", "বাংলা পাঁচ"],
                    ["6", "", "বাংলা ছয়"],
                ],
            )
            write_workbook(
                output_path,
                ["ID", "Keyword", "Success", "Response"],
                [
                    ["1", keyword("1", "বাংলা এক"), "Yes", response("পরিষ্কার বাংলা")],
                    ["2", keyword("2", "বাংলা দুই"), "Yes", "Contains SEO phrase"],
                    ["3", keyword("3", "বাংলা তিন"), "Yes", "Contains keyword phrase"],
                    ["4", keyword("4", "বাংলা চার"), "Yes", "Contains LSI phrase"],
                    ["5", keyword("5", "বাংলা পাঁচ"), "Yes", "এখানে ইনপুটে আছে"],
                    ["6", keyword("6", "বাংলা ছয়"), "Yes", "এখানে ইনপুট আছে"],
                ],
            )

            summaries = hadith.reconcile_folders(
                input_dir,
                output_dir,
                cleaned_dir,
                missing_input_dir=missing_dir,
                dry_run=False,
            )

            cleaned_rows = read_rows(cleaned_dir / "BN_BUKHARI.xlsx")
            missing_rows = read_rows(missing_dir / "[V1] BUKHARI-BANGLA_missing.xlsx")
            self.assertEqual(summaries[0].kept_rows, 1)
            self.assertEqual(summaries[0].blocked_response_rows, 5)
            self.assertEqual([row[0] for row in cleaned_rows[1:]], ["1"])
            self.assertEqual([row[0] for row in missing_rows[1:]], ["2", "3", "4", "5", "6"])

    def test_wrong_language_response_rows_are_filtered_by_language(self):
        cases = [
            ("BN", "english heading", True),
            ("BN", "اردو سرخی", True),
            ("BN", "বাংলা حَدِيث", False),
            ("EN", "বাংলা শিরোনাম", True),
            ("EN", "اردو سرخی", True),
            ("UR", "english heading", True),
            ("UR", "বাংলা শিরোনাম", True),
            ("UR", "اردو حَدِيث", False),
        ]
        for language, content_text, should_remove in cases:
            with self.subTest(language=language, content_text=content_text):
                pair = hadith.WorkbookPair(
                    language=language,
                    collection="BUKHARI",
                    input_path=Path(f"{language}_BUKHARI_INPUT.xlsx"),
                    target_path=Path(f"{language}_BUKHARI_OUTPUT.xlsx"),
                )
                with tempfile.TemporaryDirectory() as tmp:
                    target = Path(tmp) / pair.target_path.name
                    output = Path(tmp) / "cleaned.xlsx"
                    pair = hadith.WorkbookPair(pair.language, pair.collection, pair.input_path, target)
                    triple = hadith.HadithTriple("1", "", "বাংলা অনুবাদ")
                    write_workbook(
                        target,
                        ["ID", "Keyword", "Success", "Response"],
                        [["1", keyword("1"), "Yes", response(content_text)]],
                    )

                    summary = hadith.filter_target_workbook(pair, {triple}, output)[0]
                    cleaned_rows = read_rows(output)

                    self.assertEqual(summary.wrong_language_rows, 1 if should_remove else 0)
                    self.assertEqual(len(cleaned_rows) - 1, 0 if should_remove else 1)

    def test_invalid_json_response_is_not_removed_by_language_filter(self):
        pair = hadith.WorkbookPair(
            language="BN",
            collection="BUKHARI",
            input_path=Path("BN_BUKHARI_INPUT.xlsx"),
            target_path=Path("BN_BUKHARI_OUTPUT.xlsx"),
        )
        with tempfile.TemporaryDirectory() as tmp:
            target = Path(tmp) / pair.target_path.name
            output = Path(tmp) / "cleaned.xlsx"
            pair = hadith.WorkbookPair(pair.language, pair.collection, pair.input_path, target)
            triple = hadith.HadithTriple("1", "", "বাংলা অনুবাদ")
            write_workbook(
                target,
                ["ID", "Keyword", "Success", "Response"],
                [["1", keyword("1"), "Yes", "english heading but not json"]],
            )

            summary = hadith.filter_target_workbook(pair, {triple}, output)[0]
            cleaned_rows = read_rows(output)

            self.assertEqual(summary.wrong_language_rows, 0)
            self.assertEqual(len(cleaned_rows) - 1, 1)


if __name__ == "__main__":
    unittest.main()
