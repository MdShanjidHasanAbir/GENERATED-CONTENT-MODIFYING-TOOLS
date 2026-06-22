from __future__ import annotations

import argparse
import csv
import json
import re
import sys
import unicodedata
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook


LANGUAGE_IDS = ("EN", "BN", "UR", "ID")
COLLECTION_TOKENS = ("BUKHARI", "DAWUD", "MAJAH", "MUSLIM", "NASAI", "TIRMIDHI")
BOOK_ALIASES = {
    "BUKHARI": ("BUKHARI",),
    "DAWUD": ("ABU DAWUD", "DAWUD"),
    "MAJAH": ("IBN MAJAH", "MAJAH"),
    "MUSLIM": ("MUSLIM",),
    "NASAI": ("NASAI",),
    "TIRMIDHI": ("TIRMIDHI",),
}
CATALOG_BOOK_ALIASES = {
    "BULUGUL MARAM": ("BULUGH AL MARAM", "BULUGUL MARAM"),
    "LULUWAL MARJAN": ("AL LU LU WAL MARJAN", "LULUWAL MARJAN"),
    "MISHKATUL MASABIH": ("MISKAT AL MASABIH", "MISKATUL MASABIH", "MISHKATUL MASABIH"),
    "RIYADUS SALIHIN": ("RIYADUS SALIHIN", "RIYAZUS SALEHIN"),
}
FIELD_RE = re.compile(r"\b(id|arabic|translation)\s*:", re.IGNORECASE)
DIGIT_TRANSLATION = str.maketrans("০১২৩৪৫৬৭۸۹۰۱۲۳۴۵۶۷۸۹", "01234567890123456789")
BLOCKED_RESPONSE_TERMS = ("seo", "keyword", "lsi", "ইনপুটে", "ইনপুট")
BENGALI_RE = re.compile(r"[\u0980-\u09FF]")
LATIN_RE = re.compile(r"[A-Za-z]")
URDU_PERSIAN_RE = re.compile(r"[\u0679\u067E\u0686\u0688\u0691\u0698\u06A9\u06AF\u06BA\u06BE\u06C1\u06CC\u06D2]")
RELATED_BOOK_LABELS = {
    "BUKHARI": "Sahih Bukhari",
    "MUSLIM": "Sahih Muslim",
    "NASAI": "Sunan an-Nasai",
    "DAWUD": "Sunan Abu Dawud",
    "TIRMIDHI": "Jami at-Tirmidhi",
    "MAJAH": "Sunan Ibn Majah",
    "MUWATTA MALIK": "Muwatta Malik",
    "RIYADUS SALIHIN": "Riyadus Salihin",
    "BULUGUL MARAM": "Bulugh al-Maram",
    "MISHKATUL MASABIH": "Mishkat al-Masabih",
    "SILSILA SAHIHA": "Silsila Sahiha",
    "SHAMAYELE TIRMIDHI": "Shamayele Tirmidhi",
    "TARGIB WATTAHRIB": "Targib Wattahrib",
    "LULUWAL MARJAN": "Luluwal Marjan",
}
RELATED_BOOK_ALIASES = {
    "BUKHARI": (
        "bukhari",
        "al bukhari",
        "al-bukhari",
        "sahih bukhari",
        "sahih al bukhari",
        "sahih al-bukhari",
        "সহীহ বুখারী",
        "সহিহ বুখারী",
        "সহীহ বুখারি",
        "সহিহ বুখারি",
        "صحيح البخاري",
        "صحیح بخاری",
    ),
    "MUSLIM": (
        "muslim",
        "sahih muslim",
        "সহীহ মুসলিম",
        "সহিহ মুসলিম",
        "صحيح مسلم",
        "صحیح مسلم",
    ),
    "NASAI": (
        "nasai",
        "nasa i",
        "an nasai",
        "an-nasai",
        "an nasa i",
        "an-nasa i",
        "an nasa'i",
        "an-nasa'i",
        "sunan nasai",
        "sunan an nasai",
        "sunan an-nasai",
        "sunan an nasa i",
        "sunan an-nasa i",
        "sunan an nasa'i",
        "sunan an-nasa'i",
        "নাসাঈ",
        "নাসায়ী",
        "আন নাসায়ী",
        "আন-নাসায়ী",
        "সুনান আন নাসায়ী",
        "সুনান আন-নাসায়ী",
        "سنن نسائی",
    ),
    "DAWUD": (
        "abu dawud",
        "abu dawood",
        "dawud",
        "dawood",
        "sunan abu dawud",
        "sunan abu dawood",
        "সুনান আবূ দাউদ",
        "সুনানে আবূ দাউদ",
        "সুনান আবু দাউদ",
        "সুনানে আবু দাউদ",
        "আবু দাউদ",
        "سنن ابی داود",
        "سنن ابو داود",
    ),
    "TIRMIDHI": (
        "tirmidhi",
        "tirmizi",
        "jami at tirmidhi",
        "jami at-tirmidhi",
        "sunan tirmidhi",
        "তিরমিযী",
        "তিরমিজি",
        "জামে আত তিরমিজী",
        "জামে আত-তিরমিজী",
        "সুনান আত তিরমিজী",
        "সুনানে আত-তিরমিজী",
        "سنن ترمذی",
        "جامع ترمذی",
    ),
    "MAJAH": (
        "ibn majah",
        "ibn majah",
        "majah",
        "sunan ibn majah",
        "ইবনে মাজাহ",
        "ইবনু মাজাহ",
        "সুনানে ইবনু মাজাহ",
        "سنن ابن ماجہ",
        "ابن ماجہ",
    ),
    "MUWATTA MALIK": ("muwatta", "muwatta malik", "malik", "মুয়াত্তা"),
    "RIYADUS SALIHIN": (
        "riyadus salihin",
        "riyazus salehin",
        "riyad",
        "riyaz",
        "রিয়াদুস সলেহিন",
        "রিয়াদুস সালেহীন",
        "রিয়াদুস সলিহীন",
        "রিয়াযুস সালেহীন",
    ),
    "BULUGUL MARAM": (
        "bulugh al maram",
        "bulugul maram",
        "bulugh",
        "bulugul",
        "বুলুগুল মারাম",
        "বুলূগুল মারাম",
    ),
    "MISHKATUL MASABIH": (
        "mishkatul masabih",
        "mishkat al masabih",
        "miskatul masabih",
        "mishkat",
        "miskat",
        "মিশকাতুল মাসাবীহ",
        "মিশকাতুল মাসাবিহ",
    ),
    "SILSILA SAHIHA": ("silsila sahiha", "সিলসিলা সহিহা", "সিলসিলা সহীহা"),
    "SHAMAYELE TIRMIDHI": (
        "shamayele tirmidhi",
        "shamail tirmidhi",
        "শামায়েলে তিরমিযি",
        "শামায়েলে তিরমিজি",
    ),
    "TARGIB WATTAHRIB": (
        "targib wattahrib",
        "targib wat tahrib",
        "তারগিব ওয়াত তাহরিব",
        "সহিহ তারগিব ওয়াত তাহরিব",
    ),
    "LULUWAL MARJAN": (
        "luluwal marjan",
        "lu lu wal marjan",
        "লুলুওয়াল মারজান",
        "আল লু লু ওয়াল মারজান",
        "আল লু'লু ওয়াল মারজান",
    ),
}


@dataclass(frozen=True)
class HadithTriple:
    id: str
    arabic: str
    translation: str


@dataclass(frozen=True)
class WorkbookPair:
    language: str
    collection: str
    input_path: Path
    target_path: Path


@dataclass(frozen=True)
class SheetSummary:
    language: str
    collection: str
    input_file: str
    target_file: str
    output_file: str
    sheet_name: str
    original_rows: int
    kept_rows: int
    deleted_rows: int
    parse_failures: int
    success_no_rows: int
    duplicate_keyword_rows: int
    blocked_response_rows: int = 0
    wrong_language_rows: int = 0


@dataclass(frozen=True)
class MissingId:
    language: str
    collection: str
    missing_id: str
    input_file: str
    cleaned_file: str


@dataclass(frozen=True)
class BookInfo:
    book_id: str
    language_id: str
    book_name: str
    collection: str


def normalize_id(value) -> str:
    text = normalize_text(value).translate(DIGIT_TRANSLATION)
    if not text:
        return ""
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return text
    if number == number.to_integral_value():
        return str(number.quantize(Decimal(1)))
    return format(number.normalize(), "f")


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value))
    return " ".join(text.replace("\u00a0", " ").split())


def parse_keyword(keyword) -> HadithTriple | None:
    text = normalize_text(keyword)
    if not text:
        return None

    matches = list(FIELD_RE.finditer(text))
    if not matches:
        return None

    fields = {}
    for index, match in enumerate(matches):
        name = match.group(1).lower()
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(text)
        fields.setdefault(name, normalize_text(text[start:end]))

    if not all(fields.get(name) for name in ("id", "translation")):
        return None

    return HadithTriple(
        id=normalize_id(fields["id"]),
        arabic=normalize_text(fields.get("arabic")),
        translation=normalize_text(fields["translation"]),
    )


def pair_workbooks(
    input_paths: Iterable[Path],
    target_paths: Iterable[Path],
    books: dict[tuple[str, str], BookInfo] | None = None,
) -> list[WorkbookPair]:
    inputs_by_token = _index_by_workbook_key(input_paths, books)
    targets_by_token = _index_by_workbook_key(target_paths, books)
    pairs = []
    for key in sorted(set(inputs_by_token) & set(targets_by_token)):
        language, token = key
        input_path = inputs_by_token.get(key)
        target_path = targets_by_token.get(key)
        if input_path and target_path:
            pairs.append(WorkbookPair(language, token, input_path, target_path))
    return pairs


def load_input_triples(input_path: Path) -> set[HadithTriple]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    triples: set[HadithTriple] = set()
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            if not {"id", "arabic", "translation"}.issubset(columns):
                continue

            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                triple = HadithTriple(
                    id=normalize_id(_cell(row, columns["id"])),
                    arabic=normalize_text(_cell(row, columns["arabic"])),
                    translation=normalize_text(_cell(row, columns["translation"])),
                )
                if triple.id and triple.translation:
                    triples.add(triple)
    finally:
        workbook.close()
    return triples


def reconcile_folders(
    input_dir: Path | str = Path("INPUT"),
    target_dir: Path | str = Path("OUTPUT"),
    cleaned_dir: Path | str = Path("reconciled_output"),
    missing_input_dir: Path | str = Path("missing_input"),
    books_path: Path | str = Path("BOOKS.xlsx"),
    final_dir: Path | str = Path("BOOK WISE FINAL"),
    updated_content_dir: Path | str = Path("BOOK WISE FINAL UPDATED CONTENT"),
    dry_run: bool = False,
) -> list[SheetSummary]:
    input_dir = Path(input_dir)
    target_dir = Path(target_dir)
    cleaned_dir = Path(cleaned_dir)
    missing_input_dir = Path(missing_input_dir)
    books_path = Path(books_path)
    final_dir = Path(final_dir)
    updated_content_dir = Path(updated_content_dir)

    input_paths = discover_workbooks(input_dir)
    target_paths = discover_workbooks(target_dir)
    books = load_books_catalog(books_path) if books_path.exists() else None
    pairs = pair_workbooks(input_paths, target_paths, books)

    if not dry_run:
        cleaned_dir.mkdir(parents=True, exist_ok=True)

    summaries: list[SheetSummary] = []
    for pair in pairs:
        input_triples = load_input_triples(pair.input_path)
        output_path = cleaned_dir / pair.target_path.name
        pair_summaries = filter_target_workbook(pair, input_triples, output_path, dry_run=dry_run)
        summaries.extend(pair_summaries)

    if not dry_run:
        write_summary_csv(cleaned_dir / "reconciliation_summary.csv", summaries)
        missing_ids = write_missing_id_reports(input_dir, cleaned_dir, books)
        write_missing_input_workbooks(input_dir, missing_input_dir, missing_ids)
        if books_path.exists():
            write_book_wise_final(cleaned_dir, books_path, final_dir, books)
            write_updated_content_workbooks(final_dir, updated_content_dir)

    return summaries


def filter_target_workbook(
    pair: WorkbookPair,
    input_triples: set[HadithTriple],
    output_path: Path,
    dry_run: bool = False,
) -> list[SheetSummary]:
    workbook = load_workbook(pair.target_path, read_only=True, data_only=False)
    output_workbook = None if dry_run else Workbook(write_only=True)

    summaries: list[SheetSummary] = []
    seen_keyword_triples: set[HadithTriple] = set()
    try:
        for sheet in workbook.worksheets:
            output_sheet = None
            if output_workbook is not None:
                output_sheet = output_workbook.create_sheet(title=sheet.title)

            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header and output_sheet is not None:
                output_sheet.append(list(header))

            columns = _header_map(header)
            if "keyword" not in columns:
                if output_sheet is not None:
                    for row in rows:
                        output_sheet.append(list(row))
                continue

            original_rows = 0
            kept_rows = 0
            parse_failures = 0
            success_no_rows = 0
            duplicate_keyword_rows = 0
            blocked_response_rows = 0
            wrong_language_rows = 0

            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue

                original_rows += 1
                keyword = _cell(row, columns["keyword"])
                triple = parse_keyword(keyword)
                success_is_no = _success_is_no(row, columns)
                response = _cell(row, columns["response"]) if "response" in columns else None
                if success_is_no:
                    success_no_rows += 1

                if triple is None:
                    parse_failures += 1
                elif success_is_no:
                    continue
                elif triple not in input_triples:
                    continue
                elif response_has_blocked_terms(response):
                    blocked_response_rows += 1
                    continue
                elif response_has_wrong_language(response, pair.language):
                    wrong_language_rows += 1
                    continue
                elif triple in seen_keyword_triples:
                    duplicate_keyword_rows += 1
                    continue
                else:
                    seen_keyword_triples.add(triple)
                    kept_rows += 1
                    if output_sheet is not None:
                        output_sheet.append(list(row))

            summaries.append(
                SheetSummary(
                    language=pair.language,
                    collection=pair.collection,
                    input_file=pair.input_path.name,
                    target_file=pair.target_path.name,
                    output_file=output_path.name,
                    sheet_name=sheet.title,
                    original_rows=original_rows,
                    kept_rows=kept_rows,
                    deleted_rows=original_rows - kept_rows,
                    parse_failures=parse_failures,
                    success_no_rows=success_no_rows,
                    duplicate_keyword_rows=duplicate_keyword_rows,
                    blocked_response_rows=blocked_response_rows,
                    wrong_language_rows=wrong_language_rows,
                )
            )

        if output_workbook is not None:
            output_workbook.save(output_path)
    finally:
        workbook.close()

    return summaries


def response_has_blocked_terms(value) -> bool:
    text = normalize_text(value).casefold()
    if not text:
        return False
    return any(term.casefold() in text for term in BLOCKED_RESPONSE_TERMS)


def response_has_wrong_language(value, language: str) -> bool:
    text_parts = _response_content_language_text(value)
    if not text_parts:
        return False
    text = " ".join(text_parts)
    language = normalize_text(language).upper()
    has_bengali = bool(BENGALI_RE.search(text))
    has_latin = bool(LATIN_RE.search(text))
    has_urdu_persian = bool(URDU_PERSIAN_RE.search(text))
    if language == "BN":
        return has_latin or has_urdu_persian
    if language == "EN":
        return has_bengali or has_urdu_persian
    if language == "UR":
        return has_latin or has_bengali
    return False


def _response_content_language_text(value) -> list[str]:
    data = _parse_final_content_json(normalize_text(value))
    if not isinstance(data, dict):
        return []

    fields = []
    nested_content = data.get("content") if isinstance(data.get("content"), dict) else {}
    fields.extend(_text_values_from_json_value(nested_content.get("heading")))
    fields.extend(_text_values_from_json_value(nested_content.get("explanation")))
    fields.extend(_text_values_from_json_value(data.get("heading")))
    fields.extend(_text_values_from_json_value(data.get("explanation")))
    return [text for text in fields if text]


def _text_values_from_json_value(value) -> list[str]:
    if value is None:
        return []
    if isinstance(value, str):
        text = normalize_text(value)
        return [text] if text else []
    if isinstance(value, (int, float)):
        return [normalize_text(value)]
    if isinstance(value, dict):
        texts = []
        for child in value.values():
            texts.extend(_text_values_from_json_value(child))
        return texts
    if isinstance(value, list):
        texts = []
        for child in value:
            texts.extend(_text_values_from_json_value(child))
        return texts
    return []


def write_summary_csv(path: Path, summaries: Iterable[SheetSummary]) -> None:
    fieldnames = [
        "language",
        "collection",
        "input_file",
        "target_file",
        "output_file",
        "sheet_name",
        "original_rows",
        "kept_rows",
        "deleted_rows",
        "parse_failures",
        "success_no_rows",
        "duplicate_keyword_rows",
        "blocked_response_rows",
        "wrong_language_rows",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for summary in summaries:
            writer.writerow({name: getattr(summary, name) for name in fieldnames})


def write_missing_id_reports(
    input_dir: Path,
    cleaned_dir: Path,
    books: dict[tuple[str, str], BookInfo] | None = None,
) -> list[MissingId]:
    pairs = pair_workbooks(discover_workbooks(input_dir), discover_workbooks(cleaned_dir), books)
    missing_ids: list[MissingId] = []
    summary_rows = []

    for pair in pairs:
        input_ids = load_input_ids(pair.input_path)
        cleaned_ids = load_cleaned_ids(pair.target_path)
        missing_for_pair = [hadith_id for hadith_id in input_ids if hadith_id not in cleaned_ids]

        for hadith_id in missing_for_pair:
            missing_ids.append(
                MissingId(
                    language=pair.language,
                    collection=pair.collection,
                    missing_id=hadith_id,
                    input_file=str(pair.input_path),
                    cleaned_file=str(pair.target_path),
                )
            )

        summary_rows.append(
            {
                "language": pair.language,
                "collection": pair.collection,
                "input_id_count": len(input_ids),
                "cleaned_id_count": len(cleaned_ids),
                "missing_id_count": len(missing_for_pair),
            }
        )

    with (cleaned_dir / "missing_ids_report.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["language", "collection", "missing_id", "input_file", "cleaned_file"],
        )
        writer.writeheader()
        for item in missing_ids:
            writer.writerow(
                {
                    "language": item.language,
                    "collection": item.collection,
                    "missing_id": item.missing_id,
                    "input_file": item.input_file,
                    "cleaned_file": item.cleaned_file,
                }
            )

    with (cleaned_dir / "missing_ids_summary.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["language", "collection", "input_id_count", "cleaned_id_count", "missing_id_count"],
        )
        writer.writeheader()
        writer.writerows(summary_rows)

    grouped_lines = []
    current_key = None
    current_ids = []
    for item in missing_ids:
        item_key = (item.language, item.collection)
        if current_key is not None and item_key != current_key:
            grouped_lines.extend([f"[{current_key[0]} {current_key[1]}]", ", ".join(current_ids), ""])
            current_ids = []
        current_key = item_key
        current_ids.append(item.missing_id)
    if current_key is not None:
        grouped_lines.extend([f"[{current_key[0]} {current_key[1]}]", ", ".join(current_ids), ""])
    (cleaned_dir / "missing_ids_by_collection.txt").write_text(
        "\n".join(grouped_lines),
        encoding="utf-8",
    )

    return missing_ids


def write_missing_input_workbooks(input_dir: Path, missing_input_dir: Path, missing_ids: Iterable[MissingId]) -> None:
    missing_by_input: dict[str, set[str]] = {}
    for item in missing_ids:
        missing_by_input.setdefault(item.input_file, set()).add(item.missing_id)

    missing_input_dir.mkdir(parents=True, exist_ok=True)
    for old_file in missing_input_dir.glob("*.xlsx"):
        old_file.unlink()

    if not missing_by_input:
        return

    for input_file, ids in missing_by_input.items():
        input_path = Path(input_file)
        rows = load_input_rows_by_id(input_path)
        output_rows = [rows[hadith_id] for hadith_id in rows if hadith_id in ids]
        if not output_rows:
            continue

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "missing"
        sheet.append(["id", "arabic", "translation"])
        for row in output_rows:
            sheet.append([row.id, row.arabic, row.translation])

        output_name = f"{input_path.stem}_missing.xlsx"
        workbook.save(missing_input_dir / output_name)


def load_input_rows_by_id(input_path: Path) -> dict[str, HadithTriple]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    rows_by_id: dict[str, HadithTriple] = {}
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            if not {"id", "arabic", "translation"}.issubset(columns):
                continue
            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                triple = HadithTriple(
                    id=normalize_id(_cell(row, columns["id"])),
                    arabic=normalize_text(_cell(row, columns["arabic"])),
                    translation=normalize_text(_cell(row, columns["translation"])),
                )
                if triple.id and triple.id not in rows_by_id:
                    rows_by_id[triple.id] = triple
    finally:
        workbook.close()
    return rows_by_id


def load_input_ids(input_path: Path) -> list[str]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    ids = []
    seen = set()
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            if "id" not in columns:
                continue
            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                hadith_id = normalize_id(_cell(row, columns["id"]))
                if hadith_id and hadith_id not in seen:
                    seen.add(hadith_id)
                    ids.append(hadith_id)
    finally:
        workbook.close()
    return ids


def load_cleaned_ids(cleaned_path: Path) -> set[str]:
    workbook = load_workbook(cleaned_path, read_only=True, data_only=True)
    ids = set()
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            keyword_index = columns.get("keyword")
            id_index = columns.get("id")
            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                hadith_id = ""
                if keyword_index is not None:
                    parsed = parse_keyword(_cell(row, keyword_index))
                    if parsed:
                        hadith_id = parsed.id
                if not hadith_id and id_index is not None:
                    hadith_id = normalize_id(_cell(row, id_index))
                if hadith_id:
                    ids.add(hadith_id)
    finally:
        workbook.close()
    return ids


def write_book_wise_final(
    reconciled_dir: Path | str = Path("reconciled_output"),
    books_path: Path | str = Path("BOOKS.xlsx"),
    final_dir: Path | str = Path("BOOK WISE FINAL"),
    books: dict[tuple[str, str], BookInfo] | None = None,
    source_paths: Iterable[Path | str] | None = None,
    clear_existing: bool = True,
) -> dict[str, int]:
    reconciled_dir = Path(reconciled_dir)
    books_path = Path(books_path)
    final_dir = Path(final_dir)
    books = books or load_books_catalog(books_path)
    rows_by_language_book: dict[tuple[str, str], list[list[str]]] = {}

    for path in _source_workbook_paths(reconciled_dir, source_paths):
        key = _workbook_key(path, books)
        if not key:
            continue
        language, collection = key
        book = books.get((language, collection))
        if not book:
            continue
        rows_by_language_book.setdefault((language, book.book_name), [])
        rows_by_language_book[(language, book.book_name)].extend(_final_rows_from_reconciled_workbook(path, book))

    final_dir.mkdir(parents=True, exist_ok=True)
    if clear_existing:
        for old_file in final_dir.glob("*.xlsx"):
            old_file.unlink()
    for language in LANGUAGE_IDS:
        language_dir = final_dir / language
        language_dir.mkdir(parents=True, exist_ok=True)
        if clear_existing:
            for old_file in language_dir.glob("*.xlsx"):
                old_file.unlink()

    row_counts = {}
    for (language, book_name), rows in sorted(rows_by_language_book.items()):
        if not rows:
            continue
        rows.sort(key=_final_row_sort_key)
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(title="final")
        sheet.append(["book_id", "language_id", "hadith_id", "content"])
        for row in rows:
            sheet.append(row)
        output_name = f"{language}_{_safe_filename(book_name)}.xlsx"
        output_path = final_dir / language / output_name
        workbook.save(output_path)
        row_counts[f"{language}_{_safe_filename(book_name)}"] = len(rows)

    return row_counts


def convert_final_content_json(
    value,
    language: str | None = None,
    related_book_ids: dict[tuple[str, str], str] | None = None,
    related_book_aliases: dict[str, tuple[str, ...]] | None = None,
) -> str:
    converted, _error_details = convert_final_content_json_with_errors(
        value,
        language=language,
        related_book_ids=related_book_ids,
        related_book_aliases=related_book_aliases,
    )
    return converted


def convert_final_content_json_with_errors(
    value,
    language: str | None = None,
    related_book_ids: dict[tuple[str, str], str] | None = None,
    related_book_aliases: dict[str, tuple[str, ...]] | None = None,
) -> tuple[str, str]:
    text = normalize_text(value)
    data = _parse_final_content_json(text)
    if data is None:
        converted = _target_content_json_from_text(text)
        return (
            json.dumps(converted, ensure_ascii=False, separators=(",", ":")),
            "content: invalid JSON; original content wrapped in heading.description",
        )
    converted = _target_content_json_from_data(data)
    related_hadiths, errors = _normalize_related_hadiths(
        converted.get("related_hadiths", []),
        language=language,
        related_book_ids=related_book_ids,
        related_book_aliases=related_book_aliases,
    )
    converted["related_hadiths"] = related_hadiths
    return json.dumps(converted, ensure_ascii=False, separators=(",", ":")), "; ".join(errors)


def write_updated_content_workbooks(
    final_dir: Path | str = Path("BOOK WISE FINAL"),
    updated_dir: Path | str = Path("BOOK WISE FINAL UPDATED CONTENT"),
    books_path: Path | str = Path("BOOKS.xlsx"),
    related_book_ids: dict[tuple[str, str], str] | None = None,
    related_book_aliases: dict[str, tuple[str, ...]] | None = None,
    source_paths: Iterable[Path | str] | None = None,
    clear_existing: bool = True,
) -> dict[str, int]:
    final_dir = Path(final_dir)
    updated_dir = Path(updated_dir)
    books_path = Path(books_path)
    final_root = final_dir.resolve()
    if related_book_ids is None or related_book_aliases is None:
        if books_path.exists():
            books = load_books_catalog(books_path)
            if related_book_ids is None:
                related_book_ids = _related_book_ids_from_catalog(books)
            if related_book_aliases is None:
                related_book_aliases = _related_book_aliases_from_catalog(books)
        else:
            if related_book_ids is None:
                related_book_ids = {}
            if related_book_aliases is None:
                related_book_aliases = {}
    updated_dir.mkdir(parents=True, exist_ok=True)
    if clear_existing:
        for old_file in updated_dir.rglob("*.xlsx"):
            old_file.unlink()
    for language in LANGUAGE_IDS:
        (updated_dir / language).mkdir(parents=True, exist_ok=True)

    row_counts: dict[str, int] = {}
    for source_path in _source_workbook_paths(final_dir, source_paths):
        relative_path = source_path.resolve().relative_to(final_root)
        output_path = updated_dir / relative_path
        output_path.parent.mkdir(parents=True, exist_ok=True)
        row_counts[relative_path.as_posix()] = _write_updated_content_workbook(
            source_path,
            output_path,
            related_book_ids,
            related_book_aliases,
        )

    return row_counts


def _source_workbook_paths(base_dir: Path, source_paths: Iterable[Path | str] | None) -> list[Path]:
    if source_paths is None:
        return discover_workbooks(base_dir)

    base_root = base_dir.resolve()
    paths = []
    for source_path in source_paths:
        path = Path(source_path)
        if not path.is_absolute():
            path = base_dir / path
        resolved = path.resolve()
        try:
            resolved.relative_to(base_root)
        except ValueError as exc:
            raise ValueError(f"Workbook is outside {base_dir}: {source_path}") from exc
        if resolved.suffix.lower() != ".xlsx" or resolved.name.startswith("~$"):
            continue
        paths.append(resolved)
    return sorted(paths)


def _write_updated_content_workbook(
    source_path: Path,
    output_path: Path,
    related_book_ids: dict[tuple[str, str], str],
    related_book_aliases: dict[str, tuple[str, ...]] | None,
) -> int:
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    output_workbook = Workbook(write_only=True)
    converted_rows = 0
    try:
        for sheet in workbook.worksheets:
            output_sheet = output_workbook.create_sheet(title=sheet.title)
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            if header is None:
                continue
            columns = _header_map(header)
            content_index = columns.get("content")
            language_index = columns.get("language_id")
            error_index = columns.get("content_error_details")
            output_header = list(header)
            if content_index is not None and error_index is None:
                output_header.append("content_error_details")
                error_index = len(output_header) - 1
            output_sheet.append(output_header)
            for row in rows:
                output_row = list(row)
                if content_index is not None and content_index < len(output_row) and output_row[content_index]:
                    language = normalize_text(_cell(output_row, language_index)) if language_index is not None else None
                    converted_content, error_details = convert_final_content_json_with_errors(
                        output_row[content_index],
                        language=language,
                        related_book_ids=related_book_ids,
                        related_book_aliases=related_book_aliases,
                    )
                    output_row[content_index] = converted_content
                    if error_index is not None:
                        while len(output_row) <= error_index:
                            output_row.append(None)
                        output_row[error_index] = error_details or None
                    converted_rows += 1
                output_sheet.append(output_row)
        output_workbook.save(output_path)
    finally:
        workbook.close()
    return converted_rows


def _parse_final_content_json(text: str):
    if not text:
        return None

    candidates = [text]
    for marker in ("{", "["):
        index = text.find(marker)
        if index > 0:
            candidates.append(text[index:])

    decoder = json.JSONDecoder()
    for candidate in candidates:
        for parser in (json.loads, lambda raw: decoder.raw_decode(raw)[0]):
            try:
                data = parser(candidate)
            except json.JSONDecodeError:
                continue
            if isinstance(data, list):
                data = next((item for item in data if isinstance(item, dict)), None)
            if isinstance(data, dict):
                return data
    return None


def _is_target_content_json(data: dict) -> bool:
    return (
        set(data) == {"meta", "heading", "teachings", "related_hadiths"}
        and isinstance(data.get("meta"), dict)
        and isinstance(data.get("heading"), dict)
        and isinstance(data.get("teachings"), list)
        and isinstance(data.get("related_hadiths"), list)
    )


def _target_content_json_from_data(data: dict) -> dict:
    nested_content = data.get("content") if isinstance(data.get("content"), dict) else {}
    meta_data = data.get("meta") if isinstance(data.get("meta"), dict) else data.get("meta_data")
    if not isinstance(meta_data, dict):
        meta_data = {}

    heading_data = data.get("heading")
    if isinstance(heading_data, dict):
        heading_title = heading_data.get("title")
        heading_description = heading_data.get("description")
    else:
        heading_title = heading_data or nested_content.get("heading")
        heading_description = data.get("explanation") or nested_content.get("explanation")

    teachings = data.get("teachings", nested_content.get("teachings", []))
    if not isinstance(teachings, list):
        teachings = [teachings] if normalize_text(teachings) else []

    related_hadiths = data.get("related_hadiths", [])
    if not isinstance(related_hadiths, list):
        related_hadiths = []

    return {
        "meta": {
            "title": normalize_text(meta_data.get("title")),
            "description": normalize_text(meta_data.get("description")),
        },
        "heading": {
            "title": normalize_text(heading_title),
            "description": normalize_text(heading_description),
        },
        "teachings": [normalize_text(item) for item in teachings if normalize_text(item)],
        "related_hadiths": related_hadiths,
    }


def _target_content_json_from_text(text: str) -> dict:
    return {
        "meta": {
            "title": "",
            "description": "",
        },
        "heading": {
            "title": "",
            "description": text,
        },
        "teachings": [],
        "related_hadiths": [],
    }


def _normalize_related_hadiths(
    related_hadiths,
    language: str | None,
    related_book_ids: dict[tuple[str, str], str] | None,
    related_book_aliases: dict[str, tuple[str, ...]] | None = None,
) -> tuple[list[dict[str, str]], list[str]]:
    if not isinstance(related_hadiths, list):
        return [], []

    language_id = normalize_text(language).lower()
    related_book_ids = related_book_ids or {}
    related_book_aliases = related_book_aliases or {}
    normalized_items: list[dict[str, str]] = []
    errors: list[str] = []

    for index, item in enumerate(related_hadiths):
        if isinstance(item, dict):
            normalized = _normalize_related_hadith_object(
                item,
                language_id=language_id,
                related_book_ids=related_book_ids,
                related_book_aliases=related_book_aliases,
            )
            if normalized:
                normalized_items.append(normalized)
            else:
                errors.append(f"related_hadiths[{index}]: invalid related hadith object")
            continue

        normalized, error = _normalize_related_hadith_text(
            item,
            index=index,
            language_id=language_id,
            related_book_ids=related_book_ids,
            related_book_aliases=related_book_aliases,
        )
        if normalized:
            normalized_items.append(normalized)
        if error:
            errors.append(error)

    return normalized_items, errors


def _normalize_related_hadith_object(
    item: dict,
    language_id: str,
    related_book_ids: dict[tuple[str, str], str],
    related_book_aliases: dict[str, tuple[str, ...]],
) -> dict[str, str] | None:
    hadith_id = normalize_id(item.get("id"))
    book_id = normalize_id(item.get("book_id"))
    label = normalize_text(item.get("label"))
    if not book_id and label:
        collection = _related_collection_from_text(
            _normalize_related_reference_text(label),
            related_book_aliases=related_book_aliases,
        )
        if collection:
            book_id = related_book_ids.get((language_id, collection)) or related_book_ids.get(("", collection))
            label = _related_book_label(collection)
    if not (hadith_id and book_id and label):
        return None
    return {"id": hadith_id, "book_id": book_id, "label": label}


def _normalize_related_hadith_text(
    item,
    index: int,
    language_id: str,
    related_book_ids: dict[tuple[str, str], str],
    related_book_aliases: dict[str, tuple[str, ...]],
) -> tuple[dict[str, str] | None, str | None]:
    original = normalize_text(item)
    searchable = _normalize_related_reference_text(original)
    collection = _related_collection_from_text(searchable, related_book_aliases=related_book_aliases)
    id_source = _text_without_related_collection_alias(searchable, collection, related_book_aliases)
    hadith_match = re.search(r"\d+[a-z]?", id_source)
    if not hadith_match:
        return None, f'related_hadiths[{index}]: missing hadith id from "{original}"'

    if not collection:
        return None, f'related_hadiths[{index}]: unknown book label "{original}"'

    book_id = related_book_ids.get((language_id, collection)) or related_book_ids.get(("", collection))
    if not book_id:
        return None, f"related_hadiths[{index}]: no BOOKS.xlsx id for {collection} in language {language_id or 'unknown'}"

    return {
        "id": hadith_match.group(0),
        "book_id": book_id,
        "label": _related_book_label(collection),
    }, None


def _related_collection_from_text(
    text: str,
    related_book_aliases: dict[str, tuple[str, ...]] | None = None,
) -> str | None:
    for collection in _related_collection_candidates(related_book_aliases):
        for alias in _related_collection_aliases(collection, related_book_aliases):
            if alias and alias in text:
                return collection
    return None


def _text_without_related_collection_alias(
    text: str,
    collection: str | None,
    related_book_aliases: dict[str, tuple[str, ...]],
) -> str:
    if not collection:
        return text
    aliases = sorted(_related_collection_aliases(collection, related_book_aliases), key=len, reverse=True)
    for alias in aliases:
        if alias and alias in text:
            return " ".join(text.replace(alias, " ", 1).split())
    return text


def _related_collection_candidates(related_book_aliases: dict[str, tuple[str, ...]] | None) -> tuple[str, ...]:
    candidates = list(RELATED_BOOK_ALIASES)
    for collection in related_book_aliases or {}:
        if collection not in candidates:
            candidates.append(collection)
    return tuple(candidates)


def _related_collection_aliases(
    collection: str,
    related_book_aliases: dict[str, tuple[str, ...]] | None,
) -> tuple[str, ...]:
    aliases = []
    aliases.extend(RELATED_BOOK_ALIASES.get(collection, ()))
    aliases.extend((related_book_aliases or {}).get(collection, ()))
    normalized_aliases = []
    seen = set()
    for alias in aliases:
        normalized = _normalize_related_reference_text(alias)
        if normalized and normalized not in seen:
            seen.add(normalized)
            normalized_aliases.append(normalized)
    return tuple(normalized_aliases)


def _related_book_label(collection: str) -> str:
    return RELATED_BOOK_LABELS.get(collection, collection.title())


def _normalize_related_reference_text(value: str) -> str:
    text = unicodedata.normalize("NFKC", str(value)).translate(DIGIT_TRANSLATION).casefold()
    text = re.sub(r"[-_]+", " ", text)
    return " ".join(text.split())


def _related_book_ids_from_catalog(books: dict[tuple[str, str], BookInfo]) -> dict[tuple[str, str], str]:
    ids: dict[tuple[str, str], str] = {}
    for (language, collection), book in books.items():
        ids[(language.lower(), collection)] = book.book_id
    return ids


def _related_book_aliases_from_catalog(books: dict[tuple[str, str], BookInfo]) -> dict[str, tuple[str, ...]]:
    aliases_by_collection: dict[str, list[str]] = {}
    for (_language, collection), book in books.items():
        aliases = aliases_by_collection.setdefault(collection, [])
        aliases.extend((book.collection, book.book_name, *_book_aliases(book)))

    normalized: dict[str, tuple[str, ...]] = {}
    for collection, aliases in aliases_by_collection.items():
        seen = set()
        collection_aliases = []
        for alias in aliases:
            normalized_alias = _normalize_related_reference_text(alias)
            if normalized_alias and normalized_alias not in seen:
                seen.add(normalized_alias)
                collection_aliases.append(normalized_alias)
        normalized[collection] = tuple(collection_aliases)
    return normalized


def load_books_catalog(books_path: Path) -> dict[tuple[str, str], BookInfo]:
    workbook = load_workbook(books_path, read_only=True, data_only=True)
    books: dict[tuple[str, str], BookInfo] = {}
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            if not {"id", "language_id", "book_name"}.issubset(columns):
                continue
            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                language_id = normalize_text(_cell(row, columns["language_id"])).lower()
                book_name = normalize_text(_cell(row, columns["book_name"])).upper()
                collection = _collection_from_book_name(book_name)
                if not language_id or not collection:
                    continue
                key = (language_id.upper(), collection)
                books[key] = BookInfo(
                    book_id=normalize_id(_cell(row, columns["id"])),
                    language_id=language_id,
                    book_name=book_name,
                    collection=collection,
                )
    finally:
        workbook.close()
    return books


def _final_rows_from_reconciled_workbook(path: Path, book: BookInfo) -> list[list[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows_out: list[list[str]] = []
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            header = next(rows, None)
            columns = _header_map(header)
            if not {"id", "response"}.issubset(columns):
                continue
            for row in rows:
                if not row or not any(cell is not None for cell in row):
                    continue
                hadith_id = normalize_id(_cell(row, columns["id"]))
                content = normalize_text(_cell(row, columns["response"]))
                if hadith_id and content:
                    rows_out.append([book.book_id, book.language_id, hadith_id, content])
    finally:
        workbook.close()
    return rows_out


def _final_row_sort_key(row: list[str]) -> tuple[int, str]:
    language = row[1].upper()
    try:
        language_index = LANGUAGE_IDS.index(language)
    except ValueError:
        language_index = len(LANGUAGE_IDS)
    return language_index, row[2]


def print_summary(summaries: Iterable[SheetSummary]) -> None:
    for summary in summaries:
        print(
            f"{summary.language} {summary.collection} | {summary.sheet_name}: "
            f"original={summary.original_rows}, kept={summary.kept_rows}, "
            f"deleted={summary.deleted_rows}, parse_failures={summary.parse_failures}, "
            f"success_no_rows={summary.success_no_rows}, "
            f"duplicate_keyword_rows={summary.duplicate_keyword_rows}, "
            f"blocked_response_rows={summary.blocked_response_rows}, "
            f"wrong_language_rows={summary.wrong_language_rows}"
        )


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(
        description="Keep only output workbook rows whose Keyword id/arabic/translation match the paired input workbook."
    )
    parser.add_argument("--input-dir", type=Path, default=Path("INPUT"))
    parser.add_argument("--target-dir", type=Path, default=Path("OUTPUT"))
    parser.add_argument("--output-dir", type=Path, default=Path("reconciled_output"))
    parser.add_argument("--missing-input-dir", type=Path, default=Path("missing_input"))
    parser.add_argument("--books-file", type=Path, default=Path("BOOKS.xlsx"))
    parser.add_argument("--final-dir", type=Path, default=Path("BOOK WISE FINAL"))
    parser.add_argument("--updated-content-dir", type=Path, default=Path("BOOK WISE FINAL UPDATED CONTENT"))
    parser.add_argument("--dry-run", action="store_true", help="Print counts without writing cleaned workbooks.")
    subparsers = parser.add_subparsers(dest="command")

    final_parser = subparsers.add_parser(
        "write-book-wise-final",
        help="Write selected reconciled_output workbooks into BOOK WISE FINAL.",
    )
    final_parser.add_argument("--reconciled-dir", type=Path, default=Path("reconciled_output"))
    final_parser.add_argument("--books-file", type=Path, default=Path("BOOKS.xlsx"))
    final_parser.add_argument("--final-dir", type=Path, default=Path("BOOK WISE FINAL"))
    final_parser.add_argument("--files", nargs="*", default=None)

    updated_parser = subparsers.add_parser(
        "update-final-content",
        help="Write selected BOOK WISE FINAL workbooks into BOOK WISE FINAL UPDATED CONTENT.",
    )
    updated_parser.add_argument("--final-dir", type=Path, default=Path("BOOK WISE FINAL"))
    updated_parser.add_argument("--updated-content-dir", type=Path, default=Path("BOOK WISE FINAL UPDATED CONTENT"))
    updated_parser.add_argument("--books-file", type=Path, default=Path("BOOKS.xlsx"))
    updated_parser.add_argument("--files", nargs="*", default=None)
    args = parser.parse_args(argv)

    if args.command == "write-book-wise-final":
        try:
            counts = write_book_wise_final(
                args.reconciled_dir,
                args.books_file,
                args.final_dir,
                source_paths=args.files,
                clear_existing=False,
            )
        except (OSError, ValueError) as exc:
            print(f"Failed to write book-wise final workbooks: {exc}", file=sys.stderr)
            return 1
        for file_key, count in counts.items():
            print(f"{file_key}: {count} rows")
        print(f"Book-wise final workbooks written to: {args.final_dir}")
        return 0

    if args.command == "update-final-content":
        try:
            counts = write_updated_content_workbooks(
                args.final_dir,
                args.updated_content_dir,
                args.books_file,
                source_paths=args.files,
                clear_existing=False,
            )
        except (OSError, ValueError) as exc:
            print(f"Failed to update final content workbooks: {exc}", file=sys.stderr)
            return 1
        for file_key, count in counts.items():
            print(f"{file_key}: {count} rows")
        print(f"Updated content workbooks written to: {args.updated_content_dir}")
        return 0

    summaries = reconcile_folders(
        args.input_dir,
        args.target_dir,
        args.output_dir,
        missing_input_dir=args.missing_input_dir,
        books_path=args.books_file,
        final_dir=args.final_dir,
        updated_content_dir=args.updated_content_dir,
        dry_run=args.dry_run,
    )
    print_summary(summaries)

    if not summaries:
        print("No matching workbook pairs found.", file=sys.stderr)
        return 1
    if not args.dry_run:
        print(f"Cleaned workbooks and summary written to: {args.output_dir}")
        print(f"Missing input workbooks written to: {args.missing_input_dir}")
        if args.books_file.exists():
            print(f"Book-wise final workbooks written to: {args.final_dir}")
            print(f"Updated content workbooks written to: {args.updated_content_dir}")
    return 0


def discover_workbooks(directory: Path) -> list[Path]:
    return sorted(Path(directory).rglob("*.xlsx"))


def _index_by_workbook_key(
    paths: Iterable[Path],
    books: dict[tuple[str, str], BookInfo] | None = None,
) -> dict[tuple[str, str], Path]:
    indexed = {}
    for path in paths:
        key = _workbook_key(path, books)
        if key and key not in indexed:
            indexed[key] = path
    return indexed


def _workbook_key(
    path: Path,
    books: dict[tuple[str, str], BookInfo] | None = None,
) -> tuple[str, str] | None:
    language = _language_id(path)
    token = _catalog_collection_token(path, language, books) if books else None
    if token is None:
        token = _collection_token(path)
    if language and token:
        return language, token
    return None


def _language_id(path: Path) -> str | None:
    for part in (path.name, *reversed(path.parts[:-1])):
        language = _leading_language_id(part)
        if language:
            return language
    return "EN"


def _leading_language_id(value: str) -> str | None:
    text = normalize_text(value).upper()
    if len(text) < 2:
        return None
    language = text[:2]
    if language not in LANGUAGE_IDS:
        return None
    if len(text) == 2 or not text[2].isalpha():
        return language
    return None


def _collection_token(path: Path) -> str | None:
    name = _normalized_filename_for_matching(path)
    for token, aliases in BOOK_ALIASES.items():
        for alias in aliases:
            if name == alias or name.startswith(alias + " "):
                return token
    if "RIYAZUS SALEHIN" in name or "RIYADUS SALIHIN" in name:
        return "RIYADUS SALIHIN"
    return None


def _catalog_collection_token(
    path: Path,
    language: str | None,
    books: dict[tuple[str, str], BookInfo] | None,
) -> str | None:
    if not language or not books:
        return None
    name = _normalized_filename_for_matching(path)
    matches = []
    for (book_language, collection), book in books.items():
        if book_language != language:
            continue
        for alias in _book_aliases(book):
            if name == alias or name.startswith(alias + " "):
                matches.append((len(alias), collection))
    if not matches:
        return None
    return max(matches)[1]


def _book_aliases(book: BookInfo) -> tuple[str, ...]:
    aliases = [book.collection, book.book_name]
    aliases.extend(CATALOG_BOOK_ALIASES.get(book.collection, ()))
    aliases.extend(CATALOG_BOOK_ALIASES.get(book.book_name, ()))
    if book.collection in BOOK_ALIASES:
        aliases.extend(BOOK_ALIASES[book.collection])
    normalized_aliases = []
    seen = set()
    for alias in aliases:
        normalized = _normalize_book_key(alias)
        if normalized and normalized not in seen:
            seen.add(normalized)
            normalized_aliases.append(normalized)
    return tuple(normalized_aliases)


def _collection_from_book_name(book_name: str) -> str | None:
    normalized = _normalize_book_key(book_name)
    for token, aliases in BOOK_ALIASES.items():
        if normalized == token:
            return token
        for alias in aliases:
            if normalized == alias:
                return token
    if normalized in {"RIYADUS SALIHIN", "RIYAZUS SALEHIN"}:
        return "RIYADUS SALIHIN"
    return normalized or None


def _normalize_book_key(value: str) -> str:
    text = normalize_text(value).upper()
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    return " ".join(text.split())


def _normalized_filename_for_matching(path: Path) -> str:
    text = re.sub(r"^\s*\[V\d+\]\s*", "", path.stem.upper())
    text = re.sub(r"[^A-Z0-9]+", " ", text)
    text = " ".join(text.split())
    words = text.split()
    if words and words[0] in LANGUAGE_IDS:
        text = " ".join(words[1:])
    if " MARGE" in text:
        text = text.split(" MARGE", 1)[0]
    return text


def _safe_filename(value: str) -> str:
    name = re.sub(r'[<>:"/\\|?*]+', "_", normalize_text(value))
    name = name.strip(" .")
    return name or "book"


def _header_map(header) -> dict[str, int]:
    if not header:
        return {}
    return {normalize_text(value).lower(): index for index, value in enumerate(header) if normalize_text(value)}


def _cell(row, index: int):
    return row[index] if index < len(row) else None


def _success_is_no(row, columns: dict[str, int]) -> bool:
    success_index = columns.get("success")
    if success_index is None:
        return False
    return normalize_text(_cell(row, success_index)).casefold() == "no"


if __name__ == "__main__":
    raise SystemExit(main())
